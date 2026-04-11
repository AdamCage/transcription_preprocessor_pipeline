"""
Benchmark: preprocess test_audio WAVs, transcribe chunks via OpenAI-compatible STT,
then save verbose_json + XLSX (WER/CER, timings, RT metrics). Эталонные .txt необязательны:
без них колонки те же (WER/CER и эталон пустые, расхождения — «нет эталона»).

Requires: uv sync (устанавливает все зависимости, включая jiwer, openpyxl, ina).
STT server assumed at --base-url (e.g. vLLM Whisper).

Example:
  uv run python scripts/eval_test_audio.py --audio-dir test_audio --base-url http://127.0.0.1:8000
  # С авторизацией vLLM:
  uv run python scripts/eval_test_audio.py --audio-dir test_audio --base-url http://127.0.0.1:8000 --api-key sk-xxx
  # Через httpx (raw POST) вместо openai клиента:
  uv run python scripts/eval_test_audio.py --stt-backend httpx --audio-dir test_audio --base-url http://127.0.0.1:8000
  # Стерео call-center: канал 0 = call_from, канал 1 = call_to; эталоны {stem}_call_from.txt и {stem}_call_to.txt
  uv run python scripts/eval_test_audio.py --stereo-call --audio-dir stereo_wavs --base-url http://127.0.0.1:8000
  # Лог по умолчанию: eval_runs/<UTC_timestamp>/eval.log рядом с report.xlsx
  uv run python scripts/eval_test_audio.py ... --log-level DEBUG
  uv run python scripts/eval_test_audio.py ... --log-file custom.log
  uv run python scripts/eval_test_audio.py ... --no-log-file

Сообщение «Using cache found in …torch\\hub\\snakers4_silero-vad…» выводит PyTorch hub:
модель Silero VAD уже скачана и берётся из кэша — это не ошибка (часть вывода может идти через print, не через logging).
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import re
import sys
import time
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from statistics import mean, median, quantiles

import librosa
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from audio_asr_pipeline.config import (
    DEFAULT_COARSE_SEGMENTER_BACKEND,
    PipelineConfig,
    VLLMTranscribeConfig,
)
from audio_asr_pipeline.io import split_stereo_channels, write_mono_wav
from audio_asr_pipeline.models import PipelineResult
from audio_asr_pipeline.pipeline import AudioTranscriptionPipeline

log = logging.getLogger(__name__)


def _resolve_out_root(args: argparse.Namespace) -> Path:
    if args.output_dir:
        return args.output_dir.resolve()
    return Path("eval_runs") / datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")


def _setup_eval_logging(args: argparse.Namespace) -> None:
    """Console + file (unless --no-log-file); -v forces DEBUG for eval and pipeline."""
    level_name = (args.log_level or "INFO").upper()
    level = getattr(logging, level_name, logging.INFO)
    if args.verbose:
        level = logging.DEBUG

    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stderr)]
    log_path: Path | None = None
    if not args.no_log_file and args.log_file:
        log_path = Path(args.log_file).resolve()
        log_path.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(log_path, encoding="utf-8"))
    for h in handlers:
        h.setFormatter(fmt)

    logging.basicConfig(level=level, handlers=handlers, force=True)

    # Шум от зависимостей (часть torch.hub всё равно может печатать через print)
    logging.getLogger("torch").setLevel(logging.WARNING)
    logging.getLogger("httpx").setLevel(logging.DEBUG if level <= logging.DEBUG else logging.INFO)

    pkg = logging.getLogger("audio_asr_pipeline")
    if args.verbose or level <= logging.DEBUG:
        pkg.setLevel(logging.DEBUG)
    else:
        pkg.setLevel(logging.INFO)

    log.debug(
        "Logging configured: level=%s, file=%s",
        logging.getLevelName(level),
        str(log_path) if log_path else "(stderr only)",
    )


def _log_eval_banner(out_root: Path) -> None:
    log.info(
        "run context | cwd=%s | python=%s | out_root=%s",
        os.getcwd(),
        sys.version.split()[0],
        out_root.resolve(),
    )


def _normalize_for_metrics(text: str) -> str:
    t = text.strip().lower()
    t = re.sub(r"[^\w\s]", "", t, flags=re.UNICODE)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _load_reference(txt_path: Path) -> str:
    if not txt_path.is_file():
        return ""
    parts: list[str] = []
    for line in txt_path.read_text(encoding="utf-8", errors="replace").splitlines():
        s = line.strip()
        if s:
            parts.append(s)
    return " ".join(parts).strip()


def _load_channel_reference(wav_path: Path, *, side: str) -> str:
    """Reference for stereo eval: {stem}_call_from.txt / {stem}_call_to.txt next to the wav."""
    if side == "call_from":
        return _load_reference(wav_path.with_name(f"{wav_path.stem}_call_from.txt"))
    if side == "call_to":
        return _load_reference(wav_path.with_name(f"{wav_path.stem}_call_to.txt"))
    raise ValueError(side)


def _join_stereo_texts(a: str, b: str) -> str:
    return " ".join(s.strip() for s in (a, b) if s and str(s).strip())


def _discrepancy_summary(
    reference: str,
    hypothesis: str,
    *,
    max_ops: int = 18,
    max_fragment_chars: int = 90,
) -> str:
    """Краткое словесное описание расхождений для колонки отчёта."""
    ref = reference.strip()
    hyp = hypothesis.strip()
    if not ref and not hyp:
        return ""
    if not ref:
        return "нет эталона (.txt пуст или отсутствует)"
    if not hyp:
        return "гипотеза пуста"

    rw = ref.split()
    hw = hyp.split()
    if rw == hw:
        return "совпадение по словам (как в ячейках)"

    sm = SequenceMatcher(a=rw, b=hw, autojunk=False)
    lines: list[str] = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            continue
        frag_r = " ".join(rw[i1:i2])
        frag_h = " ".join(hw[j1:j2])
        if len(frag_r) > max_fragment_chars:
            frag_r = frag_r[: max_fragment_chars - 1] + "…"
        if len(frag_h) > max_fragment_chars:
            frag_h = frag_h[: max_fragment_chars - 1] + "…"
        if tag == "replace":
            lines.append(f"замена: «{frag_r}» → «{frag_h}»")
        elif tag == "delete":
            lines.append(f"в эталоне есть, в гипотезе нет: «{frag_r}»")
        elif tag == "insert":
            lines.append(f"в гипотезе лишнее: «{frag_h}»")
        if len(lines) >= max_ops:
            lines.append(f"… и ещё фрагментов (показано {max_ops})")
            break
    return "\n".join(lines)


# Заливки для основных расхождений (светлые, читаемы в Excel light/dark)
_FILL_REF = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_FILL_HYP_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_HYP_DIFF = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_DISC = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_NEUTRAL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_FILL_HYP_NO_REF = PatternFill(
    start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
)
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def _as_float(x: object) -> float | None:
    if x is None or x == "":
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _dist_summary(values: list[float]) -> tuple[float, float, float, float, int] | None:
    """mean, median, p25, p75, count."""
    nums = [float(x) for x in values if x is not None and x == x]  # skip nan
    if not nums:
        return None
    s = sorted(nums)
    n = len(s)
    m = mean(s)
    med = median(s)
    if n == 1:
        q25 = q75 = s[0]
    else:
        cuts = quantiles(s, n=4, method="inclusive")
        q25, q75 = cuts[0], cuts[2]
    return m, med, q25, q75, n


def _append_summary_sheet(
    ws2,
    *,
    summary_wer: list[float],
    summary_cer: list[float],
    audio_dur_s: list[float],
    coarse_s: list[float],
    vad_s: list[float],
    chunking_s: list[float],
    preprocess_s: list[float],
    transcription_s: list[float],
    rt_preprocess: list[float],
    rt_transcribe: list[float],
    num_chunks: list[float],
    segments_n: list[float],
) -> None:
    hdr = ["metric", "mean", "median", "p25", "p75", "n"]
    ws2.append(hdr)
    for c in ws2[1]:
        c.font = Font(bold=True)

    def row(label: str, series: list[float]) -> None:
        st = _dist_summary(series)
        if st is None:
            ws2.append([label, "", "", "", "", 0])
            return
        mu, med, q25, q75, n = st
        ws2.append([label, mu, med, q25, q75, n])

    ws2.append([])
    ws2.append(["Quality (только при наличии .txt)", "", "", "", "", ""])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    if summary_wer:
        row("WER", summary_wer)
    if summary_cer:
        row("CER", summary_cer)
    if not summary_wer and not summary_cer:
        ws2.append(["WER/CER", "—", "—", "—", "—", 0])

    ws2.append([])
    ws2.append(["Длительность аудио (файл), с", "", "", "", "", ""])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    row("audio_duration_s", audio_dur_s)

    ws2.append([])
    ws2.append(["Длительность обработки (стена), с", "", "", "", "", ""])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    row("coarse_segmentation_s", coarse_s)
    row("vad_s", vad_s)
    row("chunking_s", chunking_s)
    row("preprocess_total_s", preprocess_s)
    row("transcription_s", transcription_s)

    ws2.append([])
    ws2.append(
        ["Скорость: сек аудио на 1 с стеночного времени (выше — лучше)", "", "", "", "", ""]
    )
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    row("rt_audio_per_preprocess_s", rt_preprocess)
    row("rt_audio_per_transcription_s", rt_transcribe)

    ws2.append([])
    ws2.append(["Сегменты / чанки (шт.)", "", "", "", "", ""])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    row("num_speech_chunks", num_chunks)
    row("segments_count_merged", segments_n)


def _set_eval_column_widths(ws) -> None:
    widths = {
        1: 24,
        2: 11,
        3: 10,
        4: 8,
        5: 10,
        6: 12,
        7: 12,
        8: 12,
        9: 12,
        10: 14,
        11: 12,
        12: 54,
        13: 54,
        14: 58,
        15: 9,
        16: 9,
        17: 40,
    }
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def _set_stereo_column_widths(ws) -> None:
    for i in range(1, 26):
        ws.column_dimensions[get_column_letter(i)].width = 14 if i > 2 else 24
    for i in range(26, 32):
        ws.column_dimensions[get_column_letter(i)].width = 52
    for i in range(32, 38):
        ws.column_dimensions[get_column_letter(i)].width = 12


def _append_summary_sheet_stereo(
    ws2,
    *,
    summary_wer_from: list[float],
    summary_wer_to: list[float],
    summary_wer_overall: list[float],
    summary_cer_from: list[float],
    summary_cer_to: list[float],
    summary_cer_overall: list[float],
    audio_dur_s: list[float],
    coarse_from: list[float],
    coarse_to: list[float],
    vad_from: list[float],
    vad_to: list[float],
    chunk_from: list[float],
    chunk_to: list[float],
    preprocess_from: list[float],
    preprocess_to: list[float],
    preprocess_sum: list[float],
    transcription_from: list[float],
    transcription_to: list[float],
    transcription_sum: list[float],
    rt_preprocess_from: list[float],
    rt_preprocess_to: list[float],
    rt_transcribe_from: list[float],
    rt_transcribe_to: list[float],
    num_chunks_from: list[float],
    num_chunks_to: list[float],
    segments_from: list[float],
    segments_to: list[float],
) -> None:
    hdr = ["metric", "mean", "median", "p25", "p75", "n"]
    ws2.append(hdr)
    for c in ws2[1]:
        c.font = Font(bold=True)

    def row(label: str, series: list[float]) -> None:
        st = _dist_summary(series)
        if st is None:
            ws2.append([label, "", "", "", "", 0])
            return
        mu, med, q25, q75, n = st
        ws2.append([label, mu, med, q25, q75, n])

    ws2.append([])
    ws2.append(["Quality stereo (при наличии эталонов)", "", "", "", "", ""])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    for label, series in (
        ("WER_call_from", summary_wer_from),
        ("WER_call_to", summary_wer_to),
        ("WER_overall", summary_wer_overall),
        ("CER_call_from", summary_cer_from),
        ("CER_call_to", summary_cer_to),
        ("CER_overall", summary_cer_overall),
    ):
        if series:
            row(label, series)

    ws2.append([])
    ws2.append(["Длительность аудио (исходный стерео файл), с", "", "", "", "", ""])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    row("audio_duration_s", audio_dur_s)

    ws2.append([])
    ws2.append(["Длительность обработки по каналам (стена), с", "", "", "", "", ""])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    row("coarse_segmentation_s_call_from", coarse_from)
    row("coarse_segmentation_s_call_to", coarse_to)
    row("vad_s_call_from", vad_from)
    row("vad_s_call_to", vad_to)
    row("chunking_s_call_from", chunk_from)
    row("chunking_s_call_to", chunk_to)
    row("preprocess_total_s_call_from", preprocess_from)
    row("preprocess_total_s_call_to", preprocess_to)
    row("preprocess_total_s_sum", preprocess_sum)
    row("transcription_s_call_from", transcription_from)
    row("transcription_s_call_to", transcription_to)
    row("transcription_s_sum", transcription_sum)

    ws2.append([])
    ws2.append(["RT: сек аудио на 1 с стеночного времени", "", "", "", "", ""])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    row("rt_audio_per_preprocess_s_call_from", rt_preprocess_from)
    row("rt_audio_per_preprocess_s_call_to", rt_preprocess_to)
    row("rt_audio_per_transcription_s_call_from", rt_transcribe_from)
    row("rt_audio_per_transcription_s_call_to", rt_transcribe_to)

    ws2.append([])
    ws2.append(["Чанки / сегменты", "", "", "", "", ""])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    row("num_speech_chunks_call_from", num_chunks_from)
    row("num_speech_chunks_call_to", num_chunks_to)
    row("segments_count_call_from", segments_from)
    row("segments_count_call_to", segments_to)


def _is_pipeline_result(x: object) -> bool:
    return isinstance(x, PipelineResult)


async def _run_stereo(
    args: argparse.Namespace,
    out_root: Path,
    wavs: list[Path],
    verbose_dir: Path,
    pipeline: AudioTranscriptionPipeline,
    jiwer: object,
    cfg: PipelineConfig,
) -> int:
    stereo_wav_dir = out_root / "stereo_mono_wav"
    stereo_wav_dir.mkdir(parents=True, exist_ok=True)
    sr = cfg.target_sample_rate

    async def process_path(p: Path) -> PipelineResult:
        return await pipeline.process_file(p)

    pbar = tqdm(total=len(wavs), desc="Processing stereo WAVs", unit="file")

    async def one_stereo(wav: Path):
        t0 = time.perf_counter()
        log.info("eval stereo start | path=%s", wav.resolve())
        try:
            y, loaded_sr = librosa.load(str(wav), sr=sr, mono=False)
        except Exception as e:  # noqa: BLE001
            log.exception("stereo load failed | %s", wav)
            pbar.update(1)
            return (wav, e, None, None, 0.0)

        if y.ndim == 1:
            dur_guess = float(len(y) / loaded_sr) if y.size else 0.0
            err = ValueError("expected stereo (2 channels) for --stereo-call")
            log.warning("%s: %s", wav.name, err)
            pbar.update(1)
            return (wav, err, None, None, dur_guess)

        try:
            ch0, ch1, n_samples = split_stereo_channels(np.asarray(y))
        except Exception as e:  # noqa: BLE001
            dur_guess = float(max(y.shape) / loaded_sr) if y.size else 0.0
            log.warning("%s: not usable as stereo for --stereo-call: %s", wav.name, e)
            pbar.update(1)
            return (wav, e, None, None, dur_guess)

        duration_sec = float(n_samples / loaded_sr)
        p_from = stereo_wav_dir / f"{wav.stem}__call_from.wav"
        p_to = stereo_wav_dir / f"{wav.stem}__call_to.wav"
        write_mono_wav(p_from, ch0, loaded_sr)
        write_mono_wav(p_to, ch1, loaded_sr)

        res_from, res_to = await asyncio.gather(process_path(p_from), process_path(p_to))
        wall_s = time.perf_counter() - t0
        log.info("eval stereo done | file=%s | wall_s=%.3f", wav.name, wall_s)
        pbar.update(1)
        return (wav, None, res_from, res_to, duration_sec)

    results = await asyncio.gather(*[one_stereo(w) for w in wavs])
    pbar.close()

    log.info("Building workbook stereo (%d rows)…", len(wavs))
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "per_file"
    headers = [
        "file",
        "audio_duration_s",
        "coarse_segmentation_s_call_from",
        "coarse_segmentation_s_call_to",
        "vad_s_call_from",
        "vad_s_call_to",
        "chunking_s_call_from",
        "chunking_s_call_to",
        "preprocess_total_s_call_from",
        "preprocess_total_s_call_to",
        "preprocess_total_s_sum",
        "transcription_s_call_from",
        "transcription_s_call_to",
        "transcription_s_sum",
        "rt_audio_per_preprocess_s_call_from",
        "rt_audio_per_preprocess_s_call_to",
        "rt_audio_per_transcription_s_call_from",
        "rt_audio_per_transcription_s_call_to",
        "hypothesis_chars_call_from",
        "hypothesis_chars_call_to",
        "hypothesis_chars_sum",
        "segments_count_call_from",
        "segments_count_call_to",
        "segments_count_sum",
        "reference_call_from",
        "reference_call_to",
        "hypothesis_call_from",
        "hypothesis_call_to",
        "discrepancies_call_from",
        "discrepancies_call_to",
        "WER_call_from",
        "CER_call_from",
        "WER_call_to",
        "CER_call_to",
        "WER_overall",
        "CER_overall",
        "error",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    summary_wer_from: list[float] = []
    summary_wer_to: list[float] = []
    summary_wer_overall: list[float] = []
    summary_cer_from: list[float] = []
    summary_cer_to: list[float] = []
    summary_cer_overall: list[float] = []
    audio_dur_s: list[float] = []
    coarse_from: list[float] = []
    coarse_to: list[float] = []
    vad_from: list[float] = []
    vad_to: list[float] = []
    chunk_from: list[float] = []
    chunk_to: list[float] = []
    preprocess_from: list[float] = []
    preprocess_to: list[float] = []
    preprocess_sum: list[float] = []
    transcription_from: list[float] = []
    transcription_to: list[float] = []
    transcription_sum: list[float] = []
    rt_pf: list[float] = []
    rt_pt: list[float] = []
    rt_tf: list[float] = []
    rt_tt: list[float] = []
    num_cf: list[float] = []
    num_ct: list[float] = []
    seg_f: list[float] = []
    seg_t: list[float] = []

    wrap_cols_1based = (25, 26, 27, 28, 29, 30)

    for item in results:
        wav, early_err, res_from, res_to, duration_sec = item
        err_parts: list[str] = []
        if early_err is not None:
            err_parts.append(f"{type(early_err).__name__}: {early_err}")

        ref_f = _load_channel_reference(wav, side="call_from")
        ref_t = _load_channel_reference(wav, side="call_to")

        st_f: dict = {}
        st_t: dict = {}
        hyp_f = ""
        hyp_t = ""
        segs_f_n = 0
        segs_t_n = 0

        def consume(res: PipelineResult | None, side: str) -> None:
            nonlocal hyp_f, hyp_t, segs_f_n, segs_t_n, st_f, st_t
            if res is None:
                return
            if res.error:
                err_parts.append(f"{side}: {res.error}")
            hyp = res.text or ""
            segs = res.verbose_json.get("segments") or []
            st = res.stats
            out_name = f"{wav.stem}_{side}_verbose.json"
            verbose_dir.joinpath(out_name).write_text(
                json.dumps(res.verbose_json, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            if side == "call_from":
                hyp_f = hyp
                segs_f_n = len(segs)
                st_f = st
            else:
                hyp_t = hyp
                segs_t_n = len(segs)
                st_t = st

        if early_err is None:
            consume(res_from, "call_from")
            consume(res_to, "call_to")

        err = "; ".join(err_parts) if err_parts else ""

        wer_f = wer_t = wer_o = cer_f = cer_t = cer_o = None
        hyp_nf = _normalize_for_metrics(hyp_f)
        hyp_nt = _normalize_for_metrics(hyp_t)
        if ref_f.strip():
            rf = _normalize_for_metrics(ref_f)
            if rf and hyp_nf:
                wer_f = float(jiwer.wer(rf, hyp_nf))
                cer_f = float(jiwer.cer(rf, hyp_nf))
                summary_wer_from.append(wer_f)
                summary_cer_from.append(cer_f)
            elif rf:
                wer_f = 1.0
                cer_f = 1.0
                summary_wer_from.append(wer_f)
                summary_cer_from.append(cer_f)
        if ref_t.strip():
            rt = _normalize_for_metrics(ref_t)
            if rt and hyp_nt:
                wer_t = float(jiwer.wer(rt, hyp_nt))
                cer_t = float(jiwer.cer(rt, hyp_nt))
                summary_wer_to.append(wer_t)
                summary_cer_to.append(cer_t)
            elif rt:
                wer_t = 1.0
                cer_t = 1.0
                summary_wer_to.append(wer_t)
                summary_cer_to.append(cer_t)
        ref_join = _join_stereo_texts(ref_f, ref_t)
        hyp_join = _join_stereo_texts(hyp_f, hyp_t)
        if ref_join.strip():
            rj = _normalize_for_metrics(ref_join)
            hj = _normalize_for_metrics(hyp_join)
            if rj and hj:
                wer_o = float(jiwer.wer(rj, hj))
                cer_o = float(jiwer.cer(rj, hj))
                summary_wer_overall.append(wer_o)
                summary_cer_overall.append(cer_o)
            elif rj:
                wer_o = 1.0
                cer_o = 1.0
                summary_wer_overall.append(wer_o)
                summary_cer_overall.append(cer_o)

        if early_err is None and duration_sec > 0:
            audio_dur_s.append(duration_sec)

        for st, coarse_l, vad_l, ch_l, prep_l, tr_l, rtp_l, rtt_l in (
            (st_f, coarse_from, vad_from, chunk_from, preprocess_from, transcription_from, rt_pf, rt_tf),
            (st_t, coarse_to, vad_to, chunk_to, preprocess_to, transcription_to, rt_pt, rt_tt),
        ):
            if not st:
                continue
            for stat_key, bucket in (
                ("coarse_segmentation_sec", coarse_l),
                ("vad_sec", vad_l),
                ("chunking_sec", ch_l),
                ("preprocess_sec", prep_l),
                ("transcription_sec", tr_l),
            ):
                sv = _as_float(st.get(stat_key))
                if sv is not None:
                    bucket.append(sv)
            r1 = st.get("rt_audio_sec_per_preprocess_sec")
            if isinstance(r1, (int, float)):
                rtp_l.append(float(r1))
            r2 = st.get("rt_audio_sec_per_transcription_sec")
            if isinstance(r2, (int, float)):
                rtt_l.append(float(r2))
        if res_from and _is_pipeline_result(res_from):
            nc = res_from.verbose_json.get("pipeline_meta", {}).get("num_speech_chunks")
            if isinstance(nc, (int, float)):
                num_cf.append(float(nc))
            seg_f.append(float(len(res_from.verbose_json.get("segments") or [])))
        if res_to and _is_pipeline_result(res_to):
            nc = res_to.verbose_json.get("pipeline_meta", {}).get("num_speech_chunks")
            if isinstance(nc, (int, float)):
                num_ct.append(float(nc))
            seg_t.append(float(len(res_to.verbose_json.get("segments") or [])))

        if early_err is None:
            pf = _as_float(st_f.get("preprocess_sec")) if st_f else None
            pt = _as_float(st_t.get("preprocess_sec")) if st_t else None
            if pf is not None or pt is not None:
                preprocess_sum.append((pf or 0.0) + (pt or 0.0))
            tf_ = _as_float(st_f.get("transcription_sec")) if st_f else None
            tt_ = _as_float(st_t.get("transcription_sec")) if st_t else None
            if tf_ is not None or tt_ is not None:
                transcription_sum.append((tf_ or 0.0) + (tt_ or 0.0))

        disc_f = _discrepancy_summary(ref_f, hyp_f)
        disc_t = _discrepancy_summary(ref_t, hyp_t)

        def gv(st: dict, k: str) -> object:
            return st.get(k, "") if st else ""

        row_cells = [
            wav.name,
            duration_sec if early_err is None else "",
            gv(st_f, "coarse_segmentation_sec"),
            gv(st_t, "coarse_segmentation_sec"),
            gv(st_f, "vad_sec"),
            gv(st_t, "vad_sec"),
            gv(st_f, "chunking_sec"),
            gv(st_t, "chunking_sec"),
            gv(st_f, "preprocess_sec"),
            gv(st_t, "preprocess_sec"),
            (
                (_as_float(st_f.get("preprocess_sec")) or 0.0) + (_as_float(st_t.get("preprocess_sec")) or 0.0)
                if (st_f or st_t)
                else ""
            ),
            gv(st_f, "transcription_sec"),
            gv(st_t, "transcription_sec"),
            (
                (_as_float(st_f.get("transcription_sec")) or 0.0) + (_as_float(st_t.get("transcription_sec")) or 0.0)
                if (st_f or st_t)
                else ""
            ),
            gv(st_f, "rt_audio_sec_per_preprocess_sec"),
            gv(st_t, "rt_audio_sec_per_preprocess_sec"),
            gv(st_f, "rt_audio_sec_per_transcription_sec"),
            gv(st_t, "rt_audio_sec_per_transcription_sec"),
            len(hyp_f),
            len(hyp_t),
            len(hyp_f) + len(hyp_t),
            segs_f_n,
            segs_t_n,
            segs_f_n + segs_t_n,
            ref_f,
            ref_t,
            hyp_f,
            hyp_t,
            disc_f,
            disc_t,
            wer_f if wer_f is not None else "",
            cer_f if cer_f is not None else "",
            wer_t if wer_t is not None else "",
            cer_t if cer_t is not None else "",
            wer_o if wer_o is not None else "",
            cer_o if cer_o is not None else "",
            err,
        ]
        ws.append(row_cells)
        row_idx = ws.max_row
        for col in wrap_cols_1based:
            ws.cell(row=row_idx, column=col).alignment = _WRAP_TOP
        has_rf = bool(ref_f.strip())
        has_rt = bool(ref_t.strip())
        if has_rf:
            ws.cell(row=row_idx, column=25).fill = _FILL_REF
        else:
            ws.cell(row=row_idx, column=25).fill = _FILL_NEUTRAL
        if has_rt:
            ws.cell(row=row_idx, column=26).fill = _FILL_REF
        else:
            ws.cell(row=row_idx, column=26).fill = _FILL_NEUTRAL
        for col, has_r, hyp, ref in (
            (27, has_rf, hyp_f, ref_f),
            (28, has_rt, hyp_t, ref_t),
        ):
            if not has_r:
                ws.cell(row=row_idx, column=col).fill = _FILL_HYP_NO_REF
            else:
                mn = _normalize_for_metrics(ref) == _normalize_for_metrics(hyp)
                ws.cell(row=row_idx, column=col).fill = (
                    _FILL_HYP_OK if (hyp.strip() and mn) else _FILL_HYP_DIFF
                )
        for col, has_r, hyp, ref in (
            (29, has_rf, hyp_f, ref_f),
            (30, has_rt, hyp_t, ref_t),
        ):
            if not has_r:
                ws.cell(row=row_idx, column=col).fill = _FILL_NEUTRAL
            else:
                mn = _normalize_for_metrics(ref) == _normalize_for_metrics(hyp)
                ws.cell(row=row_idx, column=col).fill = _FILL_HYP_OK if mn else _FILL_DISC

    _set_stereo_column_widths(ws)
    ws2 = wb.create_sheet("summary")
    _append_summary_sheet_stereo(
        ws2,
        summary_wer_from=summary_wer_from,
        summary_wer_to=summary_wer_to,
        summary_wer_overall=summary_wer_overall,
        summary_cer_from=summary_cer_from,
        summary_cer_to=summary_cer_to,
        summary_cer_overall=summary_cer_overall,
        audio_dur_s=audio_dur_s,
        coarse_from=coarse_from,
        coarse_to=coarse_to,
        vad_from=vad_from,
        vad_to=vad_to,
        chunk_from=chunk_from,
        chunk_to=chunk_to,
        preprocess_from=preprocess_from,
        preprocess_to=preprocess_to,
        preprocess_sum=preprocess_sum,
        transcription_from=transcription_from,
        transcription_to=transcription_to,
        transcription_sum=transcription_sum,
        rt_preprocess_from=rt_pf,
        rt_preprocess_to=rt_pt,
        rt_transcribe_from=rt_tf,
        rt_transcribe_to=rt_tt,
        num_chunks_from=num_cf,
        num_chunks_to=num_ct,
        segments_from=seg_f,
        segments_to=seg_t,
    )
    for i in range(1, 7):
        ws2.column_dimensions[get_column_letter(i)].width = 28 if i == 1 else 14

    xlsx_path = out_root / "report.xlsx"
    try:
        wb.save(xlsx_path)
    except OSError:
        log.exception("Cannot save report to %s", xlsx_path)
        raise
    log.info("Wrote report: %s", xlsx_path.resolve())
    log.info("Wrote verbose_json dir: %s", verbose_dir.resolve())
    log.info("Wrote stereo_mono_wav dir: %s", stereo_wav_dir.resolve())
    return 0


async def _run(args: argparse.Namespace, out_root: Path) -> int:
    try:
        import jiwer  # type: ignore[import-untyped]
    except ImportError as e:
        log.error("jiwer missing: %s", e)
        raise SystemExit("Install eval extras: uv sync --extra eval") from e

    audio_dir: Path = args.audio_dir.resolve()
    verbose_dir = out_root / "verbose_json"
    verbose_dir.mkdir(parents=True, exist_ok=True)

    wavs = sorted(audio_dir.glob("*.wav"))
    log.info(
        "eval_test_audio start | audio_dir=%s | wav_count=%d | out_root=%s",
        audio_dir,
        len(wavs),
        out_root,
    )
    if not wavs:
        log.error("No .wav files in %s", audio_dir)
        return 1

    vllm_kwargs: dict = {
        "stt_backend": args.stt_backend,
        "base_url": args.base_url.rstrip("/"),
        "api_key": args.api_key,
        "model": args.model,
        "language": args.language,
        "trust_env": args.trust_env,
    }
    if args.stt_backend == "gemma":
        vllm_kwargs["gemma_api_style"] = args.gemma_api_style

    cfg = PipelineConfig(
        max_concurrent_files=args.concurrency,
        max_concurrent_chunks=args.concurrency,
        max_in_flight_requests=max(8, args.concurrency * 2),
        coarse_segmenter_backend=args.coarse_backend,  # type: ignore[arg-type]
        ina_force_cpu=not args.ina_allow_gpu,
        vad_backend="none" if args.no_vad else "silero",
        vllm=VLLMTranscribeConfig(**vllm_kwargs),
    )
    log.info("pipeline_config | %s", cfg.model_dump(mode="json"))

    if args.stt_backend == "gemma":
        endpoint_label = f"{args.base_url.rstrip('/')}/api/chat"
        if args.gemma_api_style == "openai_chat":
            endpoint_label = f"{args.base_url.rstrip('/')}/v1/chat/completions"
    else:
        endpoint_label = f"{args.base_url.rstrip('/')}/v1/audio/transcriptions"
    log.info(
        "STT %s | model=%s | stt_backend=%s | file_concurrency=%d | trust_env=%s",
        endpoint_label,
        args.model,
        args.stt_backend,
        args.concurrency,
        args.trust_env,
    )
    pipeline = AudioTranscriptionPipeline(cfg)

    if args.stereo_call:
        return await _run_stereo(args, out_root, wavs, verbose_dir, pipeline, jiwer, cfg)

    pbar = tqdm(total=len(wavs), desc="Processing WAVs", unit="file")

    async def one_wav(p: Path):
        t_wall0 = time.perf_counter()
        log.info("eval file start | path=%s", p.resolve())
        try:
            result = await pipeline.process_file(p)
        except Exception as e:  # noqa: BLE001
            log.exception(
                "Processing FAILED (row will be in xlsx with error): %s",
                p.resolve(),
            )
            pbar.update(1)
            return e
        wall_s = time.perf_counter() - t_wall0
        if result.error:
            log.error(
                "Processing FAILED (row will be in xlsx with error): %s | %s",
                p.resolve(),
                result.error,
            )
        else:
            log.info(
                "eval file ok | file=%s | wall_s=%.3f | chunks=%s | hyp_chars=%d | transcr_s=%s",
                p.name,
                wall_s,
                result.verbose_json.get("pipeline_meta", {}).get("num_speech_chunks"),
                len(result.text or ""),
                result.stats.get("transcription_sec"),
            )
        pbar.update(1)
        return result

    results = await asyncio.gather(*[one_wav(w) for w in wavs])
    pbar.close()

    for wav, res in zip(wavs, results, strict=True):
        if isinstance(res, Exception):
            log.warning("%s: %s: %s", wav.name, type(res).__name__, res)
        else:
            nchunks = res.verbose_json.get("pipeline_meta", {}).get("num_speech_chunks", "?")
            log.debug("%s: chunks=%s hypothesis_chars=%d", wav.name, nchunks, len(res.text or ""))

    log.info("Building workbook (%d rows)…", len(wavs))
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "per_file"
    headers = [
        "file",
        "audio_duration_s",
        "coarse_segmentation_s",
        "vad_s",
        "chunking_s",
        "preprocess_total_s",
        "transcription_s",
        "rt_audio_per_preprocess_s",
        "rt_audio_per_transcription_s",
        "hypothesis_chars",
        "segments_count",
        "reference_text",
        "hypothesis_text",
        "discrepancies",
        "WER",
        "CER",
        "error",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    summary_wer: list[float] = []
    summary_cer: list[float] = []
    audio_dur_s: list[float] = []
    coarse_s: list[float] = []
    vad_s: list[float] = []
    chunking_s: list[float] = []
    preprocess_s: list[float] = []
    transcription_s: list[float] = []
    rt_preprocess: list[float] = []
    rt_transcribe: list[float] = []
    num_chunks: list[float] = []
    segments_n: list[float] = []

    for wav, res in zip(wavs, results, strict=True):
        err = ""
        wer_v: float | None = None
        cer_v: float | None = None
        ref = _load_reference(wav.with_suffix(".txt"))
        hyp_norm = ""
        st: dict = {}

        if isinstance(res, Exception):
            err = f"{type(res).__name__}: {res}"
        else:
            st = res.stats
            hyp_raw = res.text or ""
            hyp_norm = _normalize_for_metrics(hyp_raw)
            segs = res.verbose_json.get("segments") or []
            err = res.error or ""
            if not err and not hyp_raw and segs:
                err = "hypothesis text empty (segments present; check merge/STT)"
            out_json = verbose_dir / f"{wav.stem}_verbose.json"
            out_json.write_text(
                json.dumps(res.verbose_json, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            if ref:
                ref_n = _normalize_for_metrics(ref)
                if ref_n and hyp_norm:
                    wer_v = float(jiwer.wer(ref_n, hyp_norm))
                    cer_v = float(jiwer.cer(ref_n, hyp_norm))
                    summary_wer.append(wer_v)
                    summary_cer.append(cer_v)
                elif ref_n:
                    wer_v = 1.0
                    cer_v = 1.0
                    summary_wer.append(wer_v)
                    summary_cer.append(cer_v)

            fdur = _as_float(res.verbose_json.get("duration"))
            if fdur is not None:
                audio_dur_s.append(fdur)
            for stat_key, bucket in (
                ("coarse_segmentation_sec", coarse_s),
                ("vad_sec", vad_s),
                ("chunking_sec", chunking_s),
                ("preprocess_sec", preprocess_s),
                ("transcription_sec", transcription_s),
            ):
                sv = _as_float(st.get(stat_key))
                if sv is not None:
                    bucket.append(sv)
            r1 = st.get("rt_audio_sec_per_preprocess_sec")
            if isinstance(r1, (int, float)):
                rt_preprocess.append(float(r1))
            r2 = st.get("rt_audio_sec_per_transcription_sec")
            if isinstance(r2, (int, float)):
                rt_transcribe.append(float(r2))
            nch = res.verbose_json.get("pipeline_meta", {}).get("num_speech_chunks")
            if isinstance(nch, (int, float)):
                num_chunks.append(float(nch))
            segments_n.append(float(len(segs)))

        if isinstance(res, Exception):
            ws.append([str(wav.name)] + [""] * 15 + [err])
        else:
            dur = res.verbose_json.get("duration", "")
            disc_text = _discrepancy_summary(ref, hyp_raw)
            has_ref = bool(ref.strip())
            match_norm = bool(
                has_ref
                and hyp_raw.strip()
                and _normalize_for_metrics(ref) == hyp_norm
            )
            ws.append(
                [
                    wav.name,
                    dur,
                    st.get("coarse_segmentation_sec", ""),
                    st.get("vad_sec", ""),
                    st.get("chunking_sec", ""),
                    st.get("preprocess_sec", ""),
                    st.get("transcription_sec", ""),
                    st.get("rt_audio_sec_per_preprocess_sec", ""),
                    st.get("rt_audio_sec_per_transcription_sec", ""),
                    len(hyp_raw),
                    len(segs),
                    ref,
                    hyp_raw,
                    disc_text,
                    wer_v if wer_v is not None else "",
                    cer_v if cer_v is not None else "",
                    err,
                ]
            )
            row_idx = ws.max_row
            for col in (12, 13, 14):
                ws.cell(row=row_idx, column=col).alignment = _WRAP_TOP
            if not has_ref:
                ws.cell(row=row_idx, column=12).fill = _FILL_NEUTRAL
                ws.cell(row=row_idx, column=13).fill = _FILL_HYP_NO_REF
                ws.cell(row=row_idx, column=14).fill = _FILL_NEUTRAL
            else:
                ws.cell(row=row_idx, column=12).fill = _FILL_REF
                ws.cell(row=row_idx, column=13).fill = (
                    _FILL_HYP_OK if match_norm else _FILL_HYP_DIFF
                )
                if match_norm:
                    ws.cell(row=row_idx, column=14).fill = _FILL_HYP_OK
                else:
                    ws.cell(row=row_idx, column=14).fill = _FILL_DISC

    _set_eval_column_widths(ws)
    ws2 = wb.create_sheet("summary")
    _append_summary_sheet(
        ws2,
        summary_wer=summary_wer,
        summary_cer=summary_cer,
        audio_dur_s=audio_dur_s,
        coarse_s=coarse_s,
        vad_s=vad_s,
        chunking_s=chunking_s,
        preprocess_s=preprocess_s,
        transcription_s=transcription_s,
        rt_preprocess=rt_preprocess,
        rt_transcribe=rt_transcribe,
        num_chunks=num_chunks,
        segments_n=segments_n,
    )
    for i in range(1, 7):
        ws2.column_dimensions[get_column_letter(i)].width = 28 if i == 1 else 14

    xlsx_path = out_root / "report.xlsx"
    try:
        wb.save(xlsx_path)
    except OSError:
        log.exception("Cannot save report to %s", xlsx_path)
        raise
    log.info("Wrote report: %s", xlsx_path.resolve())
    log.info("Wrote verbose_json dir: %s", verbose_dir.resolve())
    return 0


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--audio-dir", type=Path, default=Path("test_audio"))
    p.add_argument("--base-url", default="http://127.0.0.1:8000")
    p.add_argument(
        "--concurrency",
        type=int,
        default=3,
        help=(
            "Max files fully processed in parallel (each: coarse+VAD+chunking, then STT). "
            "Stereo uses two slots per source wav (one per channel)."
        ),
    )
    p.add_argument("--model", default="large-v3-turbo")
    p.add_argument("--language", default=None)
    p.add_argument("--output-dir", type=Path, default=None)
    p.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        help="DEBUG for eval + audio_asr_pipeline (includes each STT POST)",
    )
    p.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Console/file log level (-v forces DEBUG)",
    )
    p.add_argument(
        "--log-file",
        type=Path,
        default=None,
        help="Log file (UTF-8). Default: <out_root>/eval.log unless --no-log-file",
    )
    p.add_argument(
        "--no-log-file",
        action="store_true",
        help="Log to stderr only (no eval.log under out_root)",
    )
    p.add_argument(
        "--trust-env",
        action="store_true",
        help="Let httpx use HTTP(S)_PROXY from environment (default: off for local STT)",
    )
    p.add_argument(
        "--coarse-backend",
        choices=("whole_file", "ina"),
        default=DEFAULT_COARSE_SEGMENTER_BACKEND,
        help=(
            "Coarse segmenter: ina (inaSpeechSegmenter) or whole_file. "
            "Default matches PipelineConfig. ina requires installing inaSpeechSegmenter in the venv."
        ),
    )
    p.add_argument(
        "--ina-allow-gpu",
        action="store_true",
        help="Let TensorFlow use GPU for inaSpeechSegmenter (default: CPU only)",
    )
    p.add_argument(
        "--stereo-call",
        action="store_true",
        help=(
            "Stereo WAV: channel 0 = call_from, channel 1 = call_to; run pipeline per "
            "channel; refs {stem}_call_from.txt / {stem}_call_to.txt"
        ),
    )
    p.add_argument(
        "--stt-backend",
        choices=("httpx", "openai", "gemma"),
        default="openai",
        help=(
            "STT client backend: openai (recommended for vLLM Whisper, handles auth and retries natively), "
            "httpx (raw multipart POST), or gemma (Gemma-4 chat-based ASR via Ollama/vLLM). "
            "Default: openai."
        ),
    )
    p.add_argument(
        "--api-key",
        default=None,
        help="API key for STT server (sent as Authorization: Bearer). Required by some vLLM deployments.",
    )
    p.add_argument(
        "--no-vad",
        action="store_true",
        help="Skip VAD (Silero); coarse spans are chunked by duration only.",
    )
    p.add_argument(
        "--gemma-api-style",
        choices=("ollama_native", "openai_chat"),
        default="ollama_native",
        help=(
            "API style for --stt-backend gemma: ollama_native (Ollama /api/chat, current) "
            "or openai_chat (/v1/chat/completions, for vLLM). Default: ollama_native."
        ),
    )
    args = p.parse_args()

    if args.stt_backend == "gemma":
        if args.base_url == "http://127.0.0.1:8000":
            args.base_url = "http://localhost:11434"
        if args.model == "large-v3-turbo":
            args.model = "gemma4:e4b"
        if args.language is None:
            args.language = "ru"
    out_root = _resolve_out_root(args)
    out_root.mkdir(parents=True, exist_ok=True)
    if not args.no_log_file and args.log_file is None:
        args.log_file = out_root / "eval.log"
    _setup_eval_logging(args)
    _log_eval_banner(out_root)
    try:
        code = asyncio.run(_run(args, out_root))
    except KeyboardInterrupt:
        log.warning("Interrupted by user")
        raise SystemExit(130) from None
    except SystemExit:
        raise
    except Exception:
        log.exception("Fatal error in eval_test_audio (see traceback above)")
        raise SystemExit(1) from None
    raise SystemExit(code)


if __name__ == "__main__":
    main()
